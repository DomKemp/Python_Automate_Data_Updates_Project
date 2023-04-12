#Add shebang line

#! /usr/bin/python3

# Import modules
try:
    import os
    from openpyxl import load_workbook
    import csv
except ModuleNotFoundError as error:
    print(f'There was a problem importing a module: {error}')
    raise

def next_available_row(col):
    for row in range(5, ws.max_row):
        if ws.cell(row, col).value is None:
            return row
            break

def insert_transactions(row_start, transactions, col_start, col_end):
    if row_start is None:
        raise Exception('No empty row found in the spreadsheet')
    else:
        for row in enumerate(range(row_start, row_start + len(transactions))):
            for col in enumerate(range(col_start, col_end)):
                ws.cell(row[1], col[1], value = transactions[row[0]][col[0]])

income = []
expenses = []

# Open transactions.csv
with open(r'C:\Users\Domin\Desktop\Python_Automate_Data_Updates_Project\transactions.csv', newline='') as csvfile:
    transactions = csv.DictReader(csvfile, delimiter=',')
    rows = list(transactions)
    for row in rows:
        # Separate expenses from income
        if float(row['Amount'].replace(',', '')) > 0:
            income.append([row['Date'], float(row['Amount'].replace(',', '')), row['Simple Description'], row['Category']])
        else:
            expenses.append([row['Date'], float(row['Amount'].replace(',', '')), row['Simple Description'], row['Category']])

# Load Excel workbook
try:
    wb = load_workbook(filename =
r'C:\\Users\\Domin\\Desktop\\Python_Automate_Data_Updates_Project\\budget.xlsx')
except:
    print('Double check that the file is in the right folder with correct ext.')
    raise

try:
    ws = wb['Transactions']

    # Find next available rows
    exp_starting_row = next_available_row(3)
    inc_starting_row = next_available_row(8)

    #Insert transactions
    insert_transactions(exp_starting_row, expenses, 2, 6)
    insert_transactions(inc_starting_row, income, 7, 11)


# Save workbook
finally:
    wb.save(r'C:\\Users\\Domin\\Desktop\\Python_Automate_Data_Updates_Project\\budget.xlsx')
    print('Workbook saved')
#test